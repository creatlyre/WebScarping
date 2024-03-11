#!/usr/bin/env python
# coding: utf-8

# In[1]:


from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import time
import pandas as pd
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
import numpy as np
import datetime
from selenium.webdriver.common.action_chains import ActionChains
import xlsxwriter
from openpyxl import Workbook, load_workbook
import os
import shutil
import logging
import traceback

# 


# In[ ]:


# File paths
date_today = datetime.date.today().strftime("%Y-%m-%d")
output_gyg = r'output/GYG'
file_path_done =fr'output/GYG/{date_today}-DONE-GYG.csv'  
file_path_output = fr"output/GYG - {date_today}.xlsx"
link_file = fr'resource/GYG_links.csv'
avg_file = fr'resource/avg-gyg.csv'
re_run_path = fr'output/GYG/{date_today} - ReRun GYG.csv'
folder_path_with_txt_to_count_avg = 'Avg/GYG'



# In[2]:


# create logger object
logger_err = logging.getLogger('Error_logger')
logger_err.setLevel(logging.DEBUG)
logger_info = logging.getLogger('Info_logger')
logger_info.setLevel(logging.DEBUG)
logger_done = logging.getLogger('Done_logger')
logger_done.setLevel(logging.DEBUG)

# create console handler and set level to debug
ch = logging.StreamHandler()
ch.setLevel(logging.DEBUG)

# create file handler for error logs and set level to debug
fh_error = logging.FileHandler('Logs/GYG/error_logs.log')
fh_error.setLevel(logging.DEBUG)

# create file handler for info logs and set level to info
fh_info = logging.FileHandler('Logs/GYG/info_logs.log')
fh_info.setLevel(logging.INFO)

# create file handler for info logs and set level to info
fh_done = logging.FileHandler('Logs/GYG/done_logs.log')
fh_done.setLevel(logging.INFO)
# create formatter
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')

# add formatter to handlers
ch.setFormatter(formatter)
fh_error.setFormatter(formatter)
fh_info.setFormatter(formatter)
fh_done.setFormatter(formatter)

# add handlers to logger
logger_err.addHandler(ch)
logger_err.addHandler(fh_error)
logger_info.addHandler(ch)
logger_info.addHandler(fh_info)
logger_done.addHandler(ch)
logger_done.addHandler(fh_done)


# In[3]:


def handle_error_and_rerun(error):
#     recipient_error = 'wojbal3@gmail.com'
    tb = traceback.format_exc()
    logger_err.error('An error occurred: {} on {}'.format(str(error), tb))
#     subject = f'Error occurred - {time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())}'
#     message = f'<html><body><p>Error occurred: {str(error)} on {tb}</p></body></html>'
#     send_email(subject, message, recipient_error)


# In[4]:


def combine_csv_to_xlsx(date_prefix):
    # Get all CSV files with the specified date prefix
    csv_files = [file for file in os.listdir(output_gyg) if file.endswith('.csv') and file.startswith(date_prefix)]

    if not csv_files:
        print(f"No CSV files found with the date prefix '{date_prefix}'")
        return

    # Create a Pandas Excel writer using XlsxWriter as the engine
    output_file = f"{output_gyg} - {date_prefix}.xlsx"
    writer = pd.ExcelWriter(output_file, engine='xlsxwriter')

    for csv_file in csv_files:
        csv_path = os.path.join(f'{output_gyg}', csv_file)
        
        sheet_name = os.path.splitext(csv_file)[0]
        sheet_name = sheet_name.split(date_prefix + '-')[1].split('-GYG')[0]
        # Read the CSV file into a DataFrame
        df = pd.read_csv(csv_path)

        # Write the DataFrame to the Excel file
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Save the Excel file
    writer.save()
    writer.close()

    print(f"Combined CSV files with date prefix '{date_prefix}' into '{output_file}'")

    # Remove the CSV files
#     for csv_file in csv_files:
#         os.remove(csv_file)
    
    # Move the CSV files to the Archive folder
    for csv_file in csv_files:
        csv_path = os.path.join(f'{output_gyg}', csv_file)
        destination_path = os.path.join(archive_folder, csv_file)
        shutil.move(csv_path, destination_path)

    print(f"Moved {len(csv_files)} CSV file(s) to the '{archive_folder}' folder.")


# In[5]:


def daily_run_gyg():
#     link_file = fr'resource/GYG_links.csv'
    df_links = pd.read_csv(link_file)
    # t_url = df_links.iloc[15]['URL']
    # city = df_links.iloc[15]['City']
    # category = df_links.iloc[15]['Category']
    # t_url
    df_links = df_links[df_links['WhatIsIt'] == 'Category']
    # df_links = df_links.tail(95)
    EUR_City = [
        "Amsterdam", "Athens", "Barcelona", "Berlin", "Dublin", "Dubrovnik", "Florence", "Istanbul",
        "Krakow", "Lisbon", "Madrid", "Milan", "Naples", "Paris", "Porto", "Rome", "Palermo", "Venice",
        "Taormina", "Capri", "Sorrento", "Mount-Etna", "Mount-Vesuvius", "Herculaneum", "Amalfi-Coast",
        "Pompeii"
    ]

    USD_City = [
        "Las-Vegas", "New-York-City", "Cancun", "Dubai"
    ]

    GBP_City = [
        "Edinburgh", "London"
    ]
#     date_today = datetime.date.today().strftime("%Y-%m-%d")
#     file_path_done =fr'output/GYG/{date_today}-DONE-GYG.csv'  
#     file_path_output = fr"output/GYG - {date_today}.xlsx"
    if os.path.exists(file_path_output):
        print(f'Today ({date_today}) GYG done')
        return 'Done'

    if os.path.exists(file_path_done):
        done_msg = pd.read_csv(file_path_done)
        done_msg = done_msg.transpose()
        done_msg = done_msg.set_axis(done_msg.iloc[0], axis=1)
        done_msg = done_msg.iloc[1:]
        done_index = int(done_msg.iloc[0,0])
        df_links = df_links.iloc[(done_index+1):]
    else:
        print("Nothing done yet")

    # chrome_options = Options()
    # chrome_options.add_argument("--headless")
    # chrome_options.add_argument("--disable-gpu")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    driver.maximize_window()
    # Define the URL of the website we want to scrape
    start_time = time.time()
    total_pages = 0
    for index, row in df_links.iterrows():
    #     CHECK IF FILE PATH EXISIT IF SO CHECK THE DATA INSIDE

        page = 1
        max_pages = 9999
        data = []
        url_time = time.time()

        while page <= max_pages:
            url = f'{row["URL"]}&p={page}'
    #         print(url)
            if max_pages == 9999:
                max_pages = 'Set'

            driver.get(url)
            time.sleep(1)
        #     VERIFY IF THE CURRENCY IS CORRECT
            currency = driver.find_element(By.XPATH, "//a[@title='Select Currency']").text.strip()

            if row['City'] in EUR_City:
                if 'EUR' in currency:
                    pass
                else:
                    currency_switcher_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//a[@title='Select Currency']")))
                    # hover over the currency switcher button to show the menu
                    actions = ActionChains(driver)
                    actions.move_to_element(currency_switcher_button).perform()

                    # wait for the EUR currency option to be clickable
                    eur_currency_option = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//li[@class='currency-picker__item-parent item__currency item__currency--EUR']")))
                    # click on the EUR currency option to change the currency
                    eur_currency_option.click()
            elif row['City'] in USD_City:
                if 'USD' in currency:
                    pass
                else:
                    currency_switcher_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//a[@title='Select Currency']")))
                    # hover over the currency switcher button to show the menu
                    actions = ActionChains(driver)
                    actions.move_to_element(currency_switcher_button).perform()

                    # wait for the EUR currency option to be clickable
                    eur_currency_option = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//li[@class='currency-picker__item-parent item__currency item__currency--USD']")))
                    # click on the EUR currency option to change the currency
                    eur_currency_option.click()
            elif row['City'] in GBP_City:
                if 'GBP' in currency:
                    pass
                else:
                    currency_switcher_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//a[@title='Select Currency']")))
                    # hover over the currency switcher button to show the menu
                    actions = ActionChains(driver)
                    actions.move_to_element(currency_switcher_button).perform()

                    # wait for the EUR currency option to be clickable
                    eur_currency_option = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//li[@class='currency-picker__item-parent item__currency item__currency--GBP']")))
                    # click on the EUR currency option to change the currency
                    eur_currency_option.click()
            else:
                print('Missing from the list:', city)

            # Parse the HTML content of the page using Beautiful Soup
            html = driver.page_source
            soup = BeautifulSoup(html, 'html.parser')
            if max_pages == 'Set':
                try:
                    max_pages = int((soup.find('span', {'class': 'trip-item-pagination__controls-info'}).text.strip()).split(' ')[-1])        
                except:
                    max_pages = 1
                total_pages = total_pages+max_pages
    #             max_pages = 1


            # Extract the data from the HTML using Beautiful Soup
            tour_items = soup.find_all('li', {'class': 'list-element'})

            # print(tour_items)
            date_today = datetime.datetime.now().strftime('%Y-%m-%d')
            for tour_item in tour_items:
                title = tour_item.find('p', {'class': 'vertical-activity-card__title'}).text.strip()
                price = tour_item.find('div', {'class': 'baseline-pricing__value'}).text.strip()
                product_category = tour_item.find('span', {'class': 'vertical-activity-card__activity-type c-classifier-badge'}).text.strip()
                product_url = f"https://www.getyourguide.com/{tour_item.find('a')['href']}"
                position = int(tour_item['key']) + 1 + (page - 1) * 16
                siteuse = 'GYG'
                city = row['City']
                category = row['RawCategory']
                try:
                    discount = tour_item.find('div', {'class': 'baseline-pricing__value baseline-pricing__value--low'}).text.strip()
                except:
                    discount = 'N/A'
                try:
                    amount_reviews = tour_item.find('div', {'class': 'rating-overall__reviews'}).text.strip()
                except:
                    amount_reviews = 'N/A'
                try:
                    stars = tour_item.find('span', {'rating-overall__rating-number rating-overall__rating-number--right'}).text.strip()
                except:
                    stars = 'N/A'
                try:
                    booked = tour_item.find('span', {'class': 'booked-in-info-tag__badge c-marketplace-badge c-marketplace-badge--secondary'}).text.strip()
                except:
                    booked = 'N/A'
                try:
                    new_activity = tour_item.find('span', {'class': 'activity-info__badge c-marketplace-badge c-marketplace-badge--secondary'}).text.strip()
                except:
                    new_activity = 'N/A'

                text = tour_item.text.strip()

                data.append([title,product_url, price, stars, amount_reviews, discount, text, date_today, position, category, booked, siteuse, city ])


            page += 1
        url_done = time.time()
        message = f'Time for {city}-{category}: {round((url_done - url_time)/60, 3)}min | Pages: {max_pages} | AVG {round((url_done - url_time)/max_pages, 2)}s per page'
        print(message)
        logger_info.info(message)
        df = pd.DataFrame(data, columns=['Tytul', 'Tytul URL', 'Cena', 'Opinia', 'IloscOpini', 'Przecena', 'Tekst', 'Data zestawienia', 'Pozycja', 'Kategoria', 'Booked', 'SiteUse', 'Miasto'])
        df['Cena'] = df['Cena'].map(lambda x: x.split(' ')[-1])
        df['Przecena'] = df['Przecena'].map(lambda x: x.split('From')[1] if x != 'N/A' else 'N/A')
        df['IloscOpini'] = df['IloscOpini'].map(lambda x: x.split('(')[-1].split(')')[0].split(' ')[0].replace(',', '') if x != 'N/A' else x)
        df['VPN_City'] = ''
        #     display(df)

    #     add verificaiton process of adding logs if the data was collected 
    # save the dataframe in CSV or EXCEL file 

        # Assuming you have a DataFrame called df and a variable called city


    #     if category == 'Global':
    #         df.to_csv(file_path, header=True, index=False)
    #     else:
    #         df.to_csv(file_path, header=False, index=False, mode='a')
        file_path = fr'output/GYG/{date_today}-{city}-GYG.csv' 
        df.to_csv(file_path, header=not os.path.exists(file_path), index=False, mode='a')
        row.to_csv(file_path_done, header=True, index=True)    
    # mayeb mode for xlsx file for next try
        # Set the file path
    #     file_path = fr'output/{date_today}-{city}-GYG.xlsx'
        # Create a new workbook
    #     try:
    #         workbook = load_workbook(file_path)
    #     except:
    #         workbook = Workbook(fil)
        # Check if the sheet already exists
    #     if city in workbook.sheetnames:
    #         # Append the DataFrame to the existing sheet without headers
    #         with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
    #             df.to_excel(writer, sheet_name=city, startrow=writer.sheets[city].max_row, header=False, index=False)
    #     else:
    #         # Create a new sheet and write the DataFrame with headers
    #         with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    #             df.to_excel(writer, sheet_name=city, index=False, header=True)
    driver.quit()
    end_time = time.time()
    message_done = f'Done {len(df_links)} URLs in {round((end_time - start_time)/60,2)} mins | Pages: {total_pages} | AVG: {round((end_time - start_time)/total_pages, 2)}s'
    print(message_done)

    logger_done.info(message_done)
    combine_csv_to_xlsx(date_today)


# In[109]:


# Excel file which will be checked against avg values
def check_amount_data():
#     date_today = datetime.date.today().strftime("%Y-%m-%d")
#     xls = pd.ExcelFile(fr"output/GYG - 2023-05-27.xlsx")
    xls = pd.ExcelFile(fr"{output_gyg} - {date_today}.xlsx")
#     link_file = fr'resource/GYG_links.csv'
#     avg_file = fr'resource/avg-gyg.csv'
#     re_run_path = fr'output/GYG/{date_today} - ReRun GYG.csv'
    df_links = pd.read_csv(link_file)
    df_avg = pd.read_csv(avg_file)
    re_run_data = []

    city_to_get_data = df_links['City'].drop_duplicates().tolist()
    for excel_sheet_name in city_to_get_data:
    #     Check if the all excel files which are in df_links are available in created excel file
        if excel_sheet_name in xls.sheet_names:
    #         Data collected it's loaded excel file for selected city
            data_collected = xls.parse(sheet_name=excel_sheet_name)
            amount_of_data_collected = len(data_collected)
    #         print(excel_sheet_name, amount_of_data_collected)
            avg_value_city = int(df_avg[df_avg['City'] == excel_sheet_name]['Avg'])
            if abs(amount_of_data_collected - avg_value_city)/avg_value_city > 0.15 :
                print(abs(amount_of_data_collected - avg_value_city), excel_sheet_name, amount_of_data_collected, avg_value_city)
                category_to_get = df_links[(df_links['City'] == excel_sheet_name) & (df_links['WhatIsIt'] == 'Category')]['RawCategory'].tolist()
                category_collected = data_collected['Kategoria'].drop_duplicates().tolist()
    #             display(data_collected.groupby('Kategoria')['Kategoria'].count())
                for category_name in category_to_get:
                    if category_name in category_collected:
                        pass
                    else:
    #                     If the category is missing in the excel sheet add it to re-run data
                        print(f'Missing {category_name} for {excel_sheet_name}')
                        re_run_data.append([excel_sheet_name, category_name])

    #     If the excel sheet is missing add it to re-run data
        else:
            print(f'Missing {excel_sheet_name} in data')
            re_run_data.append([excel_sheet_name, 'all'])
    if len(re_run_data) > 0:
        pd.DataFrame(re_run_data).to_csv(re_run_path, index=False, header=['City', 'Category'])


# In[ ]:


def count_avg_data_required():
    import re
    import os
    # COUNT AVG PER CITY 
    # Initialize variables
    city_counts = []
    total_rows = 0
    result = []
    # Iterate over each text file in the directory
    for file_name in os.listdir(folder_path_with_txt_to_count_avg):
        if file_name.endswith('.txt'):
            file_path = os.path.join(folder_path_with_txt_to_count_avg, file_name)

            # Open the text file
            with open(file_path, 'r') as file:
                content = file.read()

                # Extract the city name using regular expressions
                city_list = re.findall(r'\d+ - ([^\n]+).', content)
                count_list = re.findall(r'\d+ rows', content)

                for item1, item2 in zip(city_list, count_list):
                    joined = str(item1) + ' ' + str(item2.split(' ')[0])
                    result.append(joined)

                for row in result:
                    city = row.split(' ')[0]

                    # Extract the row count using regular expressions
                    count_match = row.split(' ')[1]
                    count = int(count_match)
                    # Add the city and row count to the list
                    city_counts.append((city, count))

                    # Update the total row count
                    total_rows += count
    city_population = {}

    # Store population values for each city
    for city, row_count in city_counts:
        if city in city_population:
            city_population[city].append(row_count)
        else:
            city_population[city] = [row_count]

    # Calculate average population for each city
    city_avg = {}
    for city, population_list in city_population.items():
        city_avg[city] = round(sum(population_list) / len(population_list),0)

    # Print average population for each city
    #     report_str+= f"{city} - {round(avg, 0)}"
    avg_path_viator = 'resource/avg-gyg.csv'
    # with open(avg_path_viator, "w") as f:
    #                 f.write(report_str)
    df = pd.DataFrame(city_avg.items(), columns=['City', 'Avg'])
    df.to_csv(avg_path_viator, header=True, index=False)


# In[ ]:


while True:
    try:
        gyg_day = daily_run_gyg()
        if gyg_day == 'Done':
            break
        else:
            print('re-run not done yet')
    except Exception as e:
        handle_error_and_rerun(e)

# After sucessfull run check amount of data in Excel file if the data is missing collect missing city and/or categories
check_amount_data()


# In[7]:





# In[ ]:


# ADDED TO APP MY OTAS FOR CLEANING THE DATA WHEN UPLAODING
# df['Cena'] = df['Cena'].map(lambda x: x.split(x[0])[1].strip() if not x[0].isnumeric() else x)
# df['Booked'] = df['Booked'].astype(str)
# df['Przecena'] = df['Przecena'].astype(str)
# df['Booked'] = df['Booked'].map(lambda x: x.split('Booked')[1].split()[0] if len(x) > 5 else x)
# # df[df['Przecena']!='nan']['Przecena'].str.split()
# df['Przecena'] = df['Przecena'].map(lambda x: x.split()[1].replace(",", "") if len(x) > 4 else x)


# In[ ]:


# "2023-05-24 --> Done 407 URLs in 117.28 mins | Pages: 2286 | AVG: 3.08s"
# "2023-05-25 --> Done 407 URLs in 70.57 mins | Pages: 2282 | AVG: 1.86s"


# In[ ]:


# import json
# # Group the DataFrame by 'Tytul URL' and merge 'Kategoria' and 'Pozycja' into a single JSON column
# g = ddf.groupby('Tytul URL').agg({
#     'Tytul': 'first',
#     'Cena': 'first',
#     'Opinia': 'first',
#     'IloscOpini': 'first',
#     'Przecena': 'first',
#     'Tekst': 'first',
#     'Data zestawinie': 'first',
#     'Kategoria': lambda x: json.dumps(dict(zip(x, ddf.loc[x.index, 'Pozycja'].astype(str)))),
#     'SiteUse': 'first',
#     'Miasto': 'first'
# }).reset_index()


# In[ ]:


# ddf.iloc[0]['Tytul URL']
# ddf[ddf['Tytul URL'] == 'https://www.viator.com//tours/Athens/Athens-Full-Day-Private-Tour/d496-63145P3']


# In[ ]:


# import pyodbc
# server = 'sqlserver-myotas.database.windows.net'
# database = 'OTAs'
# username = 'azureadmin'
# password = 'brudnyHarry!66'   
# driver = '{ODBC Driver 18 for SQL Server}'
# cnxn = pyodbc.connect('DRIVER='+driver+';SERVER=tcp:'+server+';PORT=1433;DATABASE='+database+';UID='+username+';PWD='+ password, timeout=90)
# cursor = cnxn.cursor()
# df_krk = pd.read_sql('SELECT * FROM Krakow', cnxn)
# len(df_krk)

# df_krk.groupby('Tytul Url').agg({
#     'Tytul': 'first',
#     'Cena': 'first',
#     'Opinia': 'first',
#     'IloscOpini': 'first',
#     'Przecena': 'first',
#     'Tekst': 'first',
#     'Data zestawienia': 'first',
#     'Kategoria': lambda x: json.dumps(dict(zip(x.dropna(), ddf.loc[x.dropna().index, 'Pozycja'].astype(str)))),
#     'SiteUse': 'first',
#     'Miasto': 'first'
# }).reset_index()


# In[40]:





# In[41]:





# In[105]:





# In[108]:





# In[ ]:




