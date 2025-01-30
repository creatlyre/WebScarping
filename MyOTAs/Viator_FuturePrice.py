# %%
API_KEY = '56ed5b7f827aa5c258b3f6d3f57d36999aa949e8'


# %%

# %%
import asyncio
import nest_asyncio
from pyppeteer import connect
from datetime import datetime, timedelta
import random
import aiohttp
import sys
import pandas as pd
import os
from openpyxl import load_workbook
import time
import glob
from OTAs.logger.logger_manager import LoggerManager
from OTAs.file_management.file_path_manager_future_price import FilePathManagerFuturePrice
from OTAs.file_management.config_manager_future_price import ConfigReader
from OTAs.uploaders.azure_blob_uploader import AzureBlobUploader

# Allow nested event loops
nest_asyncio.apply()

class ViatorScraper:
    def __init__(self, api_key, url, viewer, file_manager, date_start_str, timeframe_days_to_collect=7, num_adults=1, language='en', extract_hours=False):
        self.api_key = api_key
        european_countries = [
            'al', 'ad', 'at', 'be', 'ba', 'bg', 'hr', 'cy', 'cz', 'dk', 'ee', 'fi', 
            'fr', 'de', 'gr', 'hu', 'is', 'ie', 'it', 'lv', 'li', 'lt', 'lu', 'mt', 
            'mc', 'me', 'nl', 'mk', 'no', 'pl', 'pt', 'ro', 'ru', 'sm', 'sk', 'si', 
            'es', 'se', 'ch', 'ua', 'gb']
        
        # Choose a random country code from the list for the proxy
        self.proxy_country = random.choice(european_countries)
        self.language = language
        self.date_start_str = date_start_str
        self.session_ttl_in_minutes = 15
        self.connection_url = f'wss://browser.zenrows.com?apikey={api_key}&proxy_country={self.proxy_country}&session_ttl={self.session_ttl_in_minutes}m'
        self.url = url
        self.viewer = viewer
        self.timeframe_days_to_collect = timeframe_days_to_collect
        self.num_adults = num_adults
        self.date_start = datetime.strptime(date_start_str, "%Y-%m-%d %H:00:00")
        self.extract_hours = extract_hours
        self.browser = None
        self.page = None
        self.output_filename = file_manager.output_file_path
        
        


        self.screenshot_dir = "screenshots"
        os.makedirs(self.screenshot_dir, exist_ok=True)
        self.html_output_dir = "html_snapshots"
        os.makedirs(self.html_output_dir, exist_ok=True)
        # Logger setup
        self.logger = LoggerManager(file_manager, application="future_price")
        # Load existing data if file exists
        self.load_existing_data()

        self.dates_to_collect = self.check_collected_dates(url)
        
        # CSS Selectors
        self.css_accept_button = 'button[id="_evidon-accept-button"]'
        self.css_currency_language_button = '[data-automation="language_currency_button"]'
        self.css_currency_button = '[data-automation="currency"]'
        self.css_currency_list = 'button[class*="currencyRow"]'
        self.css_date_picker = 'input[data-automation="availability-date-picker-input"]'
        self.css_calendar_overlay = 'div[data-automation="availability-date-picker-overlay"]'

        self.css_calendar_next_month_button = 'button[data-automation="availability-date-picker-calendar-next"]'
        self.css_calendar_month_current = 'div[class*="caption"]'

        self.css_calendar_elements = 'div[aria-disabled="false"]'
        self.css_price_date_element = 'div[class*="priceDate"]'
        self.css_travelers_input = 'input[placeholder="Number of travelers"]'
        self.css_decrease_button = 'button[data-automation="pax-decrease-selection"]'
        self.css_increase_button = 'button[data-automation="pax-increase-selection"]'
        self.css_apply_button = 'button[data-automation="pax-apply-button"]'
        self.css_tours_list = 'div[data-automation="tour-grade-list"]'
        self.css_tour_element = 'div[data-automation^="tour-grade-"]'
        self.css_title_element = 'span[class*="title"]'
        self.css_option_hours = 'button[data-automation="start_time"]'
        self.css_price_element = 'div[data-automation="price-breakdown-adult"]'
        self.css_buy_now_button = 'button[data-automation="tour-grade-buy-now-button"]'

        self.css_availability_search_button = 'button[data-automation="availability-search-button"]'
        self.css_tour_grade_price = 'span[data-automation="tour-grade-price"]'
        self.css_tour_title = 'h1[data-automation="product-title"]'
        self.css_product_title = '[data-automation="product-title"]'

        self.css_react_modal_not_available_for_selected_travelers = 'div[class*="ReactModal__Content--after-open"]'
        self.css_react_moda_close_button = 'button[class*="closeButton"]'

    def load_existing_data(self):
        # Define the pattern with today's date
        base_path_array = self.output_filename.rsplit('_', 5)
        pattern = f'{base_path_array[0]}_??-??-??_{base_path_array[2]}_{base_path_array[3]}_{base_path_array[4]}_{base_path_array[5]}'
        
        matching_files = glob.glob(pattern)

        if matching_files:
            # Initialize an empty list to store DataFrames
            data_frames = []
            for file in matching_files:
                try:
                    df = pd.read_excel(file, engine='openpyxl')
                    data_frames.append(df)
                    self.logger.logger_info.info(f"Loaded data from {file}")
                except Exception as e:
                    self.logger.logger_error.error(f"Error reading {file}: {e}")
            # Concatenate all DataFrames into one
            self.existing_data = pd.concat(data_frames, ignore_index=True)
        else:
            self.existing_data = pd.DataFrame()
            self.logger.logger_info.debug("No matching files found; initialized empty DataFrame.")

    def close_logger(self):
        if self.logger:
            self.logger.close()
            
    async def validate_websocket_url(self, url, timeout=10):
        try:
            async with aiohttp.ClientSession() as session:
                async with session.ws_connect(url, timeout=timeout):
                    self.logger.logger_info.info("WebSocket URL is valid and accessible.")
                    return True
        except Exception as e:
            self.logger.logger_err.error(f"WebSocket URL validation failed: {e}")
            return False

    async def connect_browser(self):
        try:
            self.browser = await asyncio.wait_for(connect(browserWSEndpoint=self.connection_url), timeout=15)
            self.logger.logger_info.info(f"Successfully connected to the remote browser. Proxy code: {self.proxy_country}")
            self.page = await self.browser.newPage()
            await self.page.setViewport({'width': 1920, 'height': 1080})
            await self.page.evaluate('''() => { document.documentElement.requestFullscreen(); }''')
            # Record the connection start time and session TTL
            self.connection_start_time = time.time()
            self.session_ttl = self.session_ttl_in_minutes * 60  # 15 minutes in seconds
        except asyncio.TimeoutError:
            self.logger.logger_err.error("Connection to the remote browser timed out. Exiting.")
            sys.exit(1)
        except Exception as e:
            self.logger.logger_err.error(f"Failed to connect to the remote browser: {e}")
            sys.exit(1)

    async def check_and_reconnect(self, url):
        remaining_time = (self.connection_start_time + self.session_ttl) - time.time()
        self.logger.logger_info.debug(f"Connection remaining time: {remaining_time}")
        if remaining_time <= 45:
            self.logger.logger_info.info("Browser session is about to expire in less than 30 seconds. Reconnecting...")
            await self.browser.close()
            await self.connect_browser()
            try:
                await self.page.goto(url, timeout=90000)
                self.logger.logger_info.info(f"Navigated to {url}")
            except asyncio.TimeoutError:
                self.logger.logger_err.error(f"Navigation to {url} timed out.")
                try:
                    exisit_date_picker = await self.page.waitForSelector(self.css_date_picker)
                    if exisit_date_picker:
                       pass
                    else:
                         # Date picker not found, take a screenshot and log HTML content
                        await self.take_top_half_screenshot(context_info="Timeout_navigation_to_URL_no_date_picker")
                        await self.log_html_content(context_info="Timeout_navigation_to_URL_no_date_picker")
                        self.logger.logger_err.error("Date picker not found; screenshot and HTML logged.")
                except:
                    await self.take_top_half_screenshot(context_info="Timeout_navigation_to_URL")
                    await self.log_html_content(context_info="Timeout_navigation_to_URL")
                    return
                
            # Handle cookies
            await self.handle_cookies()

            # Handle currency settings
            await self.handle_currency()


    async def scrape(self):
        # Check if there are dates to collect
        if not self.dates_to_collect:
            self.logger.logger_info.info(f"No dates to collect. Exiting scrape process.")
            self.logger.close_logger()
            return True
        # Validate the WebSocket URL
        is_valid_url = await self.validate_websocket_url(self.connection_url)
        if not is_valid_url:
            self.logger.logger_err.error("Invalid or inaccessible WebSocket URL. Exiting.")
            sys.exit(1)

        # Connect to the remote browser
        await self.connect_browser()

        try:
            # Go through each URL defined
            await self.process_url()


            self.logger.logger_done.info("Scraping completed successfully.")

        finally:
            # Ensure the browser is closed even if an error occurs
            await self.browser.close()
            self.logger.logger_done.info("Browser connection closed.")
            self.logger.close_logger()
            return True

    async def process_url(self):
        url = self.url
        viewer = self.viewer
        self.logger.logger_info.info(f"Processing URL: {url}, Viewer: {viewer}")

        try:
            
            await self.page.goto(url, timeout=90000)
            self.logger.logger_info.info(f"Navigated to {url}")
        except asyncio.TimeoutError:
            self.logger.logger_err.error(f"Navigation to {url} timed out.")
            await self.take_top_half_screenshot(context_info="Timeout_navigation_to_URL")
            await self.log_html_content(context_info="Timeout_navigation_to_URL")
            return

        # Handle cookies
        await self.handle_cookies()

        # Handle currency settings
        await self.handle_currency()

        # Collect data by iterating through dates and number of adults
        await self.collect_data(url, viewer)


    async def handle_cookies(self):
        try:
            accept_button = await self.page.querySelector(self.css_accept_button)
            if accept_button:
                await accept_button.click()
                self.logger.logger_info.info("Accepted cookies.")
        except Exception as e:
            self.logger.logger_err.error(f"Cookie accept button not found or failed to click: {e}")

    async def handle_currency(self):
        try:
            currency_language_button = await self.page.querySelector(self.css_currency_language_button)
            if currency_language_button:
                await currency_language_button.click()
                await asyncio.sleep(random.uniform(1, 3))
                self.logger.logger_info.info("Currency button clicked.")
                await self.page.waitForSelector(self.css_currency_button)
                currency_button = await self.page.querySelector(self.css_currency_button)
                current_currency_code = await self.page.evaluate('(element) => element.getAttribute("data-currency-code")', currency_language_button)
                self.logger.logger_info.debug(f"Current currency code is: {current_currency_code}")
                if 'EUR' not in current_currency_code:
                    await currency_button.click()
                    self.logger.logger_info.info(f"Current currency is {current_currency_code}, switching to EUR...")
                    currency_list = await self.page.querySelectorAll(self.css_currency_list)
                    for item_currency in currency_list:
                        text = await self.page.evaluate('(element) => element.textContent', item_currency)
                        if 'EUR' in text:
                            await item_currency.click()
                            await asyncio.sleep(random.uniform(1, 2))
                            self.logger.logger_info.info("Clicked on EUR button")
                            break
            else:
                self.logger.logger_err.error("Currency element not found.")
        except Exception as e:
            self.logger.logger_err.error(f"Currency element not found or failed to process: {e}")
            await self.log_html_content("Currency_element_not_found")
            await self.take_top_half_screenshot('Currency_element_not_found')
            raise e

    async def collect_data(self, url, viewer):
        # Check existing data for this URL

        for date in self.dates_to_collect:
            await self.check_and_reconnect(url)
            date_str = date.strftime("%Y-%m-%d")
            self.logger.logger_statistics.info(f"Start of processing: {date_str} extract hours set to: {self.extract_hours}")
            start_time = time.perf_counter()
            self.logger.logger_info.info(f"Processing date: {date_str}")
            

            # Initialize data lists
            prices_adults = []
            date_of_price = []
            option_of_price = []
            title_of_price = []
            url_of_price = []
            amount_of_adults = []
            data_viewer = []
            availability = []
            extraction_date = []
            title_product = []  
            language_product = []
            product_city = []
            uid_product = []
            hours_extraction = []
            # Click on the date picker
            await self.click_date_picker()
            date_found = await self.select_date_in_calendar(date, date_str)

            if not date_found:
                # If date is not available, record it and continue to next date
                self.logger.logger_info.info(f"Date {date_str} is unavailable.")
                await self.update_missing_date_spreadsheet(date_str, url, viewer)
                continue  # Skip to the next date
               
            await self.adjust_number_of_travelers()
            # Check for Modal windows for availability for travelers
            modal_exisit = await self.check_modal_prompt_not_available_day(date_str)
            if modal_exisit:
                end_time = time.perf_counter()
                self.logger.logger_statistics.info(f"End of processing: {date_str}. Time: {(end_time-start_time):.6f} hours set to: {self.extract_hours}")
                await self.update_missing_date_spreadsheet(date_str, url, viewer)
                continue # Skip to the next date
            try:
                await self.page.waitForSelector(self.css_buy_now_button, timeout=10000)
            except:
                 # Check for Modal windows for availability for travelers
                modal_exisit = await self.check_modal_prompt_not_available_day(date_str)
                if modal_exisit:
                    end_time = time.perf_counter()
                    self.logger.logger_statistics.info(f"End of processing: {date_str}. Time: {(end_time-start_time):.6f} hours set to: {self.extract_hours}")
                    await self.update_missing_date_spreadsheet(date_str, url, viewer)
                    continue # Skip to the next date

            # Two options to extract data
            search_button = await self.page.querySelector(self.css_availability_search_button)
            if search_button:
                await self.extract_tour_data(url, viewer, self.num_adults, date_str, prices_adults, date_of_price, 
                                             option_of_price, title_of_price, url_of_price, amount_of_adults, data_viewer, 
                                             availability, extraction_date, title_product, language_product, product_city, uid_product, hours_extraction)
            else:
                self.logger.logger_info.info('Scraping tour date simple')
                await self.extract_tour_date_simple(url, viewer, self.num_adults, date_str, prices_adults, date_of_price,
                                                     option_of_price, title_of_price, url_of_price, amount_of_adults, data_viewer, 
                                                     availability, extraction_date, title_product, language_product, product_city, uid_product, hours_extraction)

            # Store the collected data
            data = {
                'extraction_date': extraction_date,
                'date': date_of_price,
                'title': title_product,
                'tour_option': title_of_price,
                'time_range': option_of_price,
                'price_per_person': prices_adults,
                'language': language_product,
                'adults': amount_of_adults,
                'title_url': url_of_price,                
                'city': product_city,
                'viewer': data_viewer,
                'availability': availability,
                'uid': uid_product,
                'hours_extraction': hours_extraction
            }
            df_temp = pd.DataFrame(data)
            await self.save_data_to_excel(df_temp)
            end_time = time.perf_counter()
            self.logger.logger_statistics.info(f"End of processing: {date_str}. Time: {(end_time-start_time):.6f} hours set to: {self.extract_hours}")
            

        self.logger.logger_done.info(f"Completed processing for URL: {url}")

    def check_collected_dates(self, url):
        if not self.existing_data.empty:
            existing_url_data = self.existing_data[self.existing_data['title_url'] == url]
            existing_dates = existing_url_data['date'].unique()
        else:
            existing_dates = []

        date_max_to_do = self.date_start + timedelta(days=self.timeframe_days_to_collect)
        date_list = [self.date_start + timedelta(days=x) for x in range((date_max_to_do - self.date_start).days)]

        # Remove dates that have already been collected
        dates_to_collect = []
        for date in date_list:
            date_str = date.strftime("%Y-%m-%d")
            if date_str in existing_dates:
                self.logger.logger_info.info(f"Data for date {date_str} already collected for URL {url}. Skipping.")
            else:
                dates_to_collect.append(date)
        return dates_to_collect

    async def click_date_picker(self):
        try:
            await self.page.waitForSelector(self.css_date_picker)
            date_picker = await self.page.querySelector(self.css_date_picker)
            ### Retrieve and save the top half of the page after clicking
            # await self.take_top_half_screenshot(context_info="Before_click_date_picker")
            # await self.log_html_content(context_info="Before_click_date_picker")
            
            if date_picker:
                await date_picker.click()
                await self.page.waitForSelector(self.css_calendar_overlay)
                ## Retrieve and save the top half of the page after clicking
                # await self.take_top_half_screenshot(context_info="After_click_date_picker")
                # ## Log HTML content after clicking the date picker
                # await self.log_html_content(context_info="After_click_date_picker")


            else:
                self.logger.logger_err.error("Date picker not found.")
        except Exception as e:
            await self.take_top_half_screenshot(context_info="Failed to click date picker")
            ## Log HTML content after clicking the date picker
            await self.log_html_content(context_info="Failed to click date picker")
            raise self.logger.logger_err.error(f"Failed to click date picker: {e}")
            
            

    async def select_date_in_calendar(self, date, date_str):
        try:
            self.page.waitForSelector(self.css_calendar_overlay)
            calendar_overlay = await self.page.querySelector(self.css_calendar_overlay)
            if calendar_overlay:
                # inner_html = await self.page.evaluate('(element) => element.innerHTML', calendar_overlay)
                # self.logger.logger_info.debug(f"Calendar Overlay Inner HTML:\n{inner_html}")

                calendar_current_month = await calendar_overlay.querySelector(self.css_calendar_month_current)
                if calendar_current_month:
                    current_month_text = await self.page.evaluate('(element) => element.textContent', calendar_current_month)
                    # Check if the current month matches the required month
                    if date.strftime('%B %Y') != current_month_text.strip():
                        # Only click next button if current month is before the required month
                        if datetime.strptime(current_month_text.strip(), '%B %Y') < datetime(date.year, date.month, 1):
                            calendar_next_button = await calendar_overlay.querySelector(self.css_calendar_next_month_button)
                            if calendar_next_button:
                                await calendar_next_button.click()
                                await asyncio.sleep(1)
                                return await self.select_date_in_calendar(date, date_str)
                calendar_elements = await calendar_overlay.querySelectorAll(self.css_calendar_elements)
                date_found = False
                for day_element in calendar_elements:
                    day_element_inner = await day_element.querySelector(self.css_price_date_element)
                    if day_element_inner:
                        day_text = await self.page.evaluate('(element) => element.textContent', day_element_inner)
                        day_int = int(day_text.strip())
                        if day_int == date.day:
                            await day_element.click()
                            date_found = True
                            break
                if not date_found:
                    self.logger.logger_done.info(f"Date {date_str} not available in calendar.")
            
                return date_found
        except Exception as e:
            self.logger.logger_err.error(f"Failed to select date {date_str}: {e}")
            return False

    async def update_missing_date_spreadsheet(self, date_str, url, viewer):
            # Create DataFrame and save to CSV
            tour_title_element = await self.page.querySelector(self.css_product_title)
            tour_title = await (await tour_title_element.getProperty('textContent')).jsonValue() if tour_title_element else "Title unavailable"
            city = url.split('tours')[1].split('/')[1]
            uid = url.split('/')[-1]

            data = {
                'extraction_date': [self.date_start_str],
                'date': [date_str],
                'title': [tour_title],
                'tour_option': ['N/A'],
                'time_range': ['N/A'],
                'price_per_person': ['N/A'],
                'language': [self.language],
                'adults': [self.num_adults],
                'title_url': [url],     
                'city': [city],           
                'viewer': [viewer],
                'availability': [False],
                'uid': [uid],
                'hours_extraction': [self.extract_hours]
                
            }
            df_temp = pd.DataFrame(data)
            await self.save_data_to_excel(df_temp)
    async def adjust_number_of_travelers(self):
        try:
            travelers_input = await self.page.querySelector(self.css_travelers_input)
            if travelers_input:
                await travelers_input.click()
                num_adults_current_text = await (await travelers_input.getProperty('value')).jsonValue()
                num_adults_current = int(num_adults_current_text.split()[0])
                if num_adults_current > self.num_adults:
                    decrease_button = await self.page.querySelector(self.css_decrease_button)
                    for _ in range(num_adults_current - self.num_adults):
                        await decrease_button.click()
                        await asyncio.sleep(0.5)
                elif num_adults_current < self.num_adults:
                    increase_button = await self.page.querySelector(self.css_increase_button)
                    for _ in range(self.num_adults - num_adults_current):
                        await increase_button.click()
                        await asyncio.sleep(1)

                await self.page.waitForSelector(self.css_apply_button, timeout=5000)
                apply_button = await self.page.querySelector(self.css_apply_button)
                if apply_button:
                    await apply_button.click()
                    self.logger.logger_info.info(f"Set number of adults to {self.num_adults}")
                else:
                    self.logger.logger_err.error("Apply button not found.")
            else:
                self.logger.logger_err.error("Travelers input box not found.")
        except Exception as e:
            self.logger.logger_err.error(f"Failed to adjust number of adults to {self.num_adults}: {e}")
    
    async def check_modal_prompt_not_available_day(self, date_str):
        try:
            modal_window = await self.page.querySelector(self.css_react_modal_not_available_for_selected_travelers)
            modal_close_button = await modal_window.querySelector(self.css_react_moda_close_button)
            self.logger.logger_info.info(f'Modal windows found for date: {date_str} - Travelers: {self.num_adults}')
            # await self.take_top_half_screenshot("Modal_window_found_check")
            await self.log_html_content("Modal_window_found_check")
            if modal_close_button:
                await modal_close_button.click()
                return True
        except:
            self.logger.logger_info.debug('Modal window not appeared')
            return False

    async def has_class(self, element, class_keyword: str) -> bool:
        if element:
            try:
                classes = await self.page.evaluate(
                    '(element) => element.className', element
                )
                return class_keyword in classes.split()
            except Exception as e:
                self.logger.logger_err.error(f"Error checking class '{class_keyword}': {e}")
        return False

    async def extract_tour_data(self, url, viewer, num_adults, date_str, prices_adults, date_of_price, option_of_price, 
                                title_of_price, url_of_price, amount_of_adults, data_viewer, availability, extraction_date,
                                title_product, language_product, city_product, uid_product, hours_extraction):
        await self.page.waitForSelector(self.css_tours_list, timeout=10000)
        price_overwrite = False
        tour_title_element = await self.page.querySelector(self.css_product_title)
        tour_title = await (await tour_title_element.getProperty('textContent')).jsonValue() if tour_title_element else "Title unavailable"
        city = url.split('tours')[1].split('/')[1
        ]
        uid = url.split('/')[-1]


        try:
            tours_list = await self.page.querySelector(self.css_tours_list)
            if tours_list:
                tours = await tours_list.querySelectorAll(self.css_tour_element + ':not([data-automation="tour-grade-price"])')
                self.logger.logger_info.debug(f"Found {len(tours)} tour elements on the page.")

                for tour in tours:
                    # Try change the selected to next Tour if mulitples hours to collect
                    if self.extract_hours:
                    # Determine the state of the tour: available & selected, available & unselected, or unavailable
                        tour_is_selected = await self.has_class(tour, "selected__UlHb")
                        tour_is_unselected = await self.has_class(tour, "unselected__CIlk")
                        tour_is_unavailable = await self.has_class(tour, "unavailable__f79u")

                        if tour_is_selected:
                            self.logger.logger_info.debug("Tour is already selected. Proceeding to extract data.")
                        elif tour_is_unselected:
                            select_button = await tour.querySelector('input[type="radio"]')
                            if select_button:
                                await select_button.click()
                                self.logger.logger_info.debug("Clicked to select the unselected tour.")
                                await asyncio.sleep(0.5)  # Reduced sleep time; adjust as needed
                        elif tour_is_unavailable:
                            pass

                    try:
                        # Extract title
                        title_element = await tour.querySelector(self.css_title_element)
                        title = await (await title_element.getProperty('textContent')).jsonValue() if title_element else "Title unavailable"
                        self.logger.logger_info.debug(f"Extracted title: {title.strip()}")

                        # Extract hours options
                        hours_elements = await tour.querySelectorAll(self.css_option_hours)
                        if hours_elements and self.extract_hours:
                            self.logger.logger_info.debug(f"Found {len(hours_elements)} hours options for tour: {title.strip()}")
                            
                            for hours_button in hours_elements:
                                # Check if the hour is inactive
                                is_inactive = await self.page.evaluate(
                                    '(element) => Array.from(element.classList).some(cls => cls.includes("inactive"))', hours_button
                                )
                                # Check if the hour is already selected
                                is_selected = await self.page.evaluate(
                                    '(element) => Array.from(element.classList).some(cls => cls.includes("selected"))', hours_button
                                )

                                if is_inactive:
                                    hours_text = await (await hours_button.getProperty('textContent')).jsonValue() if hours_button else "Hours unavailable"
                                    self.logger.logger_info.debug(f"Skipping inactive hour option: {hours_text.strip()}")
                                    price_overwrite = True
                                elif is_selected:
                                    hours_text = await (await hours_button.getProperty('textContent')).jsonValue() if hours_button else "Hours unavailable"
                                    self.logger.logger_info.debug(f"Hour option already selected: {hours_text.strip()}")
                                else:
                                    # Select hour option
                                    await hours_button.click()
                                    self.logger.logger_info.debug(f"Clicked on hour option to select it.")
                                    await asyncio.sleep(1)  # Wait for page to update

                                    hours_text = await (await hours_button.getProperty('textContent')).jsonValue() if hours_button else "Hours unavailable"
                                    self.logger.logger_info.debug(f"Newly selected hour option: {hours_text.strip()}")

                                # Extract price for the selected hour
                                price_element = await tour.querySelector(self.css_price_element)
                                price = await (await price_element.getProperty('textContent')).jsonValue() if price_element else "Price unavailable"
                                if price_overwrite:
                                    price = 'Price unavailable'
                                    price_overwrite = False
                                    availability_text=False
                                else:
                                    availability_text=True

                                # Append data to lists
                                extraction_date.append(self.date_start_str)
                                prices_adults.append(price.strip())
                                date_of_price.append(date_str)
                                option_of_price.append(hours_text.strip())
                                title_of_price.append(title.strip())
                                amount_of_adults.append(num_adults)
                                url_of_price.append(url)
                                data_viewer.append(viewer)
                                availability.append(availability_text)
                                language_product.append(self.language)
                                title_product.append(tour_title)
                                city_product.append(city)
                                uid_product.append(uid)
                                hours_extraction.append(self.extract_hours)

                                self.logger.logger_info.info(f"Collected data - Date: {date_str}, Adults: {num_adults}, Title: {title.strip()}, Option: {hours_text.strip()}, Price: {price.strip()}")
                        else:
                            # Handle case when no hours are available
                            self.logger.logger_info.debug("No hours options available for this tour or turned off")
                            hours_text = 'Hours unavailable'
                            price_element = await tour.querySelector(self.css_price_element)
                            price = await (await price_element.getProperty('textContent')).jsonValue() if price_element else "Price unavailable"

                            # Append data to lists for tours without hours
                            extraction_date.append(self.date_start_str)
                            prices_adults.append(price.strip())
                            date_of_price.append(date_str)
                            option_of_price.append(hours_text.strip())
                            title_of_price.append(title.strip())
                            amount_of_adults.append(num_adults)
                            url_of_price.append(url)
                            data_viewer.append(viewer)
                            if price == 'Price unavailable':
                                availability.append(False)
                            else:
                                availability.append(True)
                            language_product.append(self.language)
                            title_product.append(tour_title)
                            city_product.append(city)
                            uid_product.append(uid)
                            hours_extraction.append(self.extract_hours)

                            self.logger.logger_info.info(f"Collected data - Date: {date_str}, Adults: {num_adults}, Title: {title.strip()}, Option: {hours_text.strip()}, Price: {price.strip()}")

                    except Exception as e:
                        self.logger.logger_err.error(f"Failed to extract data from tour '{title.strip() if title_element else 'Unknown'}': {e}")
            else:
                self.logger.logger_err.error("Tours list not found.")
        except Exception as e:
            self.logger.logger_err.error(f"Failed to extract tours: {e}")

    async def extract_tour_date_simple(self, url, viewer, num_adults, date_str, prices_adults, date_of_price, option_of_price, title_of_price, url_of_price, 
                                       amount_of_adults, data_viewer, availability, extraction_date,
                                       title_product, language_product, city_product, uid_product, hours_extraction):
        try:
            title_element = await self.page.querySelector(self.css_tour_title)
            title = await (await title_element.getProperty('textContent')).jsonValue() if title_element else "Title unavailable"
            
            price_element = await self.page.querySelector(self.css_tour_grade_price)
            price = await (await price_element.getProperty('textContent')).jsonValue() if price_element else "Price unavailable"
            
            hours_text = "Option unavailable"
            city = url.split('tours')[1].split('/')[1]
            uid = url.split('/')[-1]
            extraction_date.append(self.date_start_str)
            prices_adults.append(price.strip())
            date_of_price.append(date_str)
            option_of_price.append(hours_text.strip())
            title_product.append(title.strip())
            title_of_price.append("Option unavailable")
            amount_of_adults.append(num_adults)
            url_of_price.append(url)
            data_viewer.append(viewer)
            availability.append(True)
            language_product.append(self.language)
            city_product.append(city)
            uid_product.append(uid)
            hours_extraction.append(self.extract_hours)


            self.logger.logger_info.info(f"Collected data - Date: {date_str}, Adults: {num_adults}, Title: {title.strip()}, Option: {hours_text.strip()}, Price: {price.strip()}")

        except Exception as e:
            self.logger.logger_err.error(f"Failed to extract data from tour: {e}")
    async def save_data_to_excel(self, df_temp):
        try:
            if not os.path.exists(self.output_filename):
                # Save with headers if the file doesn't exist
                df_temp.to_excel(self.output_filename, index=False)
                self.logger.logger_info.info(f"Data saved to {self.output_filename}")
            else:
                sheet_name = 'Sheet1'  # Replace with your sheet name if different

                # Load the workbook to find the last row in the sheet
                workbook = load_workbook(self.output_filename)
                if sheet_name in workbook.sheetnames:
                    ws = workbook[sheet_name]
                    startrow = ws.max_row
                else:
                    startrow = 0

                # Use ExcelWriter to append data without setting writer.book
                with pd.ExcelWriter(self.output_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    # Write data to the Excel file starting from the calculated row
                    df_temp.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False, header=False)

                self.logger.logger_info.info(f"Data appended to {self.output_filename}")
        except Exception as e:
            self.logger.logger_err.error(f"Failed to save data to Excel: {e}")

    async def take_top_half_screenshot(self, context_info=""):
        """
        Takes a screenshot of the top half of the current page and saves it with a timestamp and context info.
        
        :param context_info: Additional information to include in the screenshot filename.
        """
        try:
            # Retrieve the viewport dimensions
            viewport = await self.page.evaluate("""() => {
                return {
                    width: window.innerWidth,
                    height: window.innerHeight
                };
            }""")
            viewport_width = viewport['width']
            viewport_height = viewport['height']
            
            # Define the clipping rectangle for the top half
            clip = {
                'x': 0,
                'y': 0,
                'width': viewport_width,
                'height': viewport_height * 2 # / 2  # Top half
            }
            
            # Generate a unique filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_context = "".join(c if c.isalnum() else "_" for c in context_info)
            filename = f"screenshot_top_half_{timestamp}_{safe_context}.png"
            filepath = os.path.join(self.screenshot_dir, filename)
            
            # Take the screenshot with the specified clip
            await self.page.screenshot({
                'path': filepath,
                'clip': clip
            })
            
            # Log the successful capture
            self.logger.logger_info.info(f"Top half screenshot saved to {filepath}")
        
        except Exception as e:
            self.logger.logger_err.error(f"Failed to take top half screenshot: {e}")


    async def log_html_content(self, context_info=""):
        """
        Retrieves the current page's HTML content and logs it.
        
        :param context_info: Additional information to include in the log entry.
        """
        try:
            # Retrieve the page's HTML content
            html_content = await self.page.content()
            
            # Option 1: Log the entire HTML content (Use with caution)
            # self.logger.logger_info.info(f"HTML Content ({context_info}):\n{html_content}")
            
            # Option 2: Save the HTML content to a file and log the file path
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_context = "".join(c if c.isalnum() else "_" for c in context_info)
            filename = f"html_snapshot_{safe_context}_{timestamp}.html"
            filepath = os.path.join(self.html_output_dir, filename)
            
            with open(filepath, "w", encoding="utf-8") as file:
                file.write(html_content)
            
            self.logger.logger_info.info(f"HTML content saved to {filepath} (Context: {context_info})")
        
        except Exception as e:
            self.logger.logger_err.error(f"Failed to log HTML content ({context_info}): {e}")

def main():

    site = 'Viator'
    file_manager__config_path = FilePathManagerFuturePrice(site, "N/A", "N/A", "N/A")
    config_reader = ConfigReader(file_manager__config_path.config_file_path)
    urls = config_reader.get_urls_by_ota(site)
    all_excel_file_list = set()
    for item in urls:
            url = item['url']
            viewer = item["viewer"]
            url_processed = False
            for config in item['configurations']:
                adults = config['adults']
                language = config['language']
                schedules = config['schedules']
                extract_hours = schedules[0].get('extract_hours')

                frequency, max_days = config_reader.get_highest_order_schedule(schedules)
                if frequency.lower() == "no schedule for today":
                    # logger_done.info(f"URL: {url} is not scheduled for today to run")
                    continue  # Use 'continue' to process other configurations
                else:
                    date_start_str = datetime.today().strftime("%Y-%m-%d %H:00:00")
                    # date_start_str = '2024-12-30 12:00:00'
                    file_manager = FilePathManagerFuturePrice(site, 'N/A', adults, language)  
                    all_excel_file_list.add((file_manager.output_file_path, file_manager.blob_name))
                    print(f"Running script for URL: {url}, Adults: {adults}, Language: {language}, Frequency: {frequency}, Max Days: {max_days}")
                    while True:
                        scraper = ViatorScraper(API_KEY, url, viewer, file_manager, date_start_str=date_start_str, timeframe_days_to_collect=max_days, 
                                                num_adults=adults,extract_hours=extract_hours)
                        url_processed = asyncio.run(scraper.scrape())
                        if url_processed:
                            break

                # Handle multiple times per day if applicable
                # config_reader_instance = config_reader  # Reference to ConfigReader instance
                # config_schedules = config.get('schedules', [])
                # for schedule in config_schedules:
                #     if schedule.get('frequency_type', '').lower() in ["twice_a_day", "three_times_a_day"]:
                #         times_per_day = schedule.get('times_per_day', 1)
                #         for time in range(times_per_day):
                #             # Implement logic for multiple runs per day
                #             # For example, trigger additional data collection
                #             print(f"Running additional collection for URL: {url}, Adults: {adults}, Language: {language}, Frequency: {schedule['frequency_type']}, Run number: {time+1}")
    ####

    azure_storage_upload = AzureBlobUploader(file_manager, LoggerManager(file_manager,'future_price'))
    print(f'Files to process: {all_excel_file_list}')
    for files_to_upload in all_excel_file_list:
        future_price_file_path = files_to_upload[0]
        future_price_blob_name = files_to_upload[1]
        
        azure_storage_upload.upload_excel_to_azure_storage_account_future_price(future_price_file_path, future_price_blob_name)
        azure_storage_upload.transform_upload_to_refined_future_price(future_price_file_path, future_price_blob_name)
# %%
if __name__ == "__main__":
    main()
# %%

# %%
### Manual upload file to Storage Account:
# file = 'Viator_2024-11-21_15-00-00_en_1_future_price.xlsx'
# file_split = file.split('_')
# manual_date = f'{file_split[1]} {file_split[2]}'
# language = file_split[3]
# adutls = file_split[4]
# file_manager = FilePathManagerFuturePrice('Viator', 'N/A', adutls, language, True, manual_date)  
# print(f'Will work on file... \n {file}')

# azure_storage_upload = AzureBlobUploader(file_manager, LoggerManagerFuturePrice(file_manager,'future_price'))
# azure_storage_upload.upload_excel_to_azure_storage_account_future_price(file_manager.output_file_path, file_manager.blob_name)
# azure_storage_upload.transform_upload_to_refined_future_price(file_manager.output_file_path, file_manager.blob_name)
# 
# %%
